from openpyxl import load_workbook

def create_pattern(paradigm):
    """Creates pattern from xlsx empirical data files"""
    if paradigm == 'face':
        file_path = 'face_data.xlsx'
    elif paradigm == 'grating':
        file_path = 'grating_data.xlsx'
    
    workbook = load_workbook(file_path)
    sheet = workbook["Sheet1"]
    pattern = {}
    current_matrix = []
    matrix_index = 0

    for row in sheet.iter_rows(values_only=True):
        if all(cell is None for cell in row):
            if current_matrix:
                pattern[matrix_index] = current_matrix
                current_matrix = []
                matrix_index += 1
                print(matrix_index)
        else:
            current_matrix.append([cell for cell in row if cell is not None])

    if current_matrix:
        pattern[matrix_index] = current_matrix
        
    return pattern
